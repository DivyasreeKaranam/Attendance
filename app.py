from flask import Flask, render_template, request, redirect, url_for, session, flash
from werkzeug.security import generate_password_hash, check_password_hash
from datetime import datetime
import openpyxl
import logging
import os

app = Flask(__name__)
app.secret_key = os.environ.get('SECRET_KEY', 'fallback_secret_key')

attendance_file = "attendance.xlsx"

# Configure logging
logging.basicConfig(level=logging.INFO)
logger = logging.getLogger(__name__)

def init_attendance_file():
    if not os.path.exists(attendance_file):
        wb = openpyxl.Workbook()
        sheet = wb.active
        sheet.title = "Attendance"
        sheet.append(["ID", "Name", "Subject", "Date", "Time"])
        wb.save(attendance_file)
        logger.info(f"Created new attendance file: {attendance_file}")

init_attendance_file()

# Sample users (for demonstration purposes)
users = {"admin": generate_password_hash("password")}

@app.route("/")
def home():
    return render_template("home.html")

@app.route("/login", methods=["GET", "POST"])
def login():
    if request.method == "POST":
        username = request.form["username"]
        password = request.form["password"]
        if username in users and check_password_hash(users[username], password):
            session["username"] = username
            flash("Logged in successfully!", "success")
            return redirect(url_for("index"))
        else:
            flash("Invalid credentials", "error")
    return render_template("login.html")

@app.route("/logout")
def logout():
    session.pop("username", None)
    flash("Logged out successfully!", "success")
    return redirect(url_for("login"))

@app.route("/index")
def index():
    if "username" in session:
        return render_template("index.html")
    else:
        flash("Please log in to access this page", "error")
        return redirect(url_for("login"))

@app.route("/take_attendance", methods=["POST"])
def take_attendance():
    if "username" in session:
        try:
            name = request.form["name"]
            subject = request.form["subject"]
            now = datetime.now()
            date = now.strftime("%Y-%m-%d")
            time = now.strftime("%H:%M:%S")

            logger.info(f"Attempting to record attendance for {name} in {subject}")

            wb = openpyxl.load_workbook(attendance_file)
            sheet = wb.active
            row_id = sheet.max_row
            sheet.append([row_id, name, subject, date, time])
            wb.save(attendance_file)

            logger.info(f"Attendance recorded successfully for {name}")
            flash("Attendance recorded successfully!", "success")
            return redirect(url_for("show_attendance"))
        except Exception as e:
            logger.error(f"Error recording attendance: {str(e)}")
            flash("An error occurred while recording attendance", "error")
            return redirect(url_for("index"))
    else:
        flash("Please log in to take attendance", "error")
        return redirect(url_for("login"))

@app.route("/attendance")
def show_attendance():
    if "username" in session:
        try:
            logger.info("Fetching attendance data")
            wb = openpyxl.load_workbook(attendance_file)
            sheet = wb.active
            attendance = []
            for row in sheet.iter_rows(min_row=2, values_only=True):
                attendance.append({
                    "id": row[0],
                    "name": row[1],
                    "subject": row[2],
                    "date": row[3],
                    "time": row[4]
                })
            logger.info(f"Fetched {len(attendance)} attendance records")
            return render_template("attendance.html", attendance=attendance)
        except Exception as e:
            logger.error(f"Error displaying attendance: {str(e)}")
            flash("An error occurred while fetching attendance data", "error")
            return redirect(url_for("index"))
    else:
        flash("Please log in to view attendance", "error")
        return redirect(url_for("login"))

@app.route("/update_attendance/<int:id>", methods=["GET", "POST"])
def update_attendance(id):
    if "username" in session:
        try:
            wb = openpyxl.load_workbook(attendance_file)
            sheet = wb.active
            
            if request.method == "POST":
                name = request.form["name"]
                subject = request.form["subject"]
                date = request.form["date"]
                time = request.form["time"]
                
                for row in sheet.iter_rows(min_row=2):
                    if row[0].value == id:
                        row[1].value = name
                        row[2].value = subject
                        row[3].value = date
                        row[4].value = time
                        break
                
                wb.save(attendance_file)
                flash("Attendance updated successfully!", "success")
                return redirect(url_for("show_attendance"))
            
            attendance = None
            for row in sheet.iter_rows(min_row=2, values_only=True):
                if row[0] == id:
                    attendance = {
                        "id": row[0],
                        "name": row[1],
                        "subject": row[2],
                        "date": row[3],
                        "time": row[4]
                    }
                    break
            
            if attendance:
                return render_template("update_attendance.html", attendance=attendance)
            else:
                flash("Attendance record not found", "error")
                return redirect(url_for("show_attendance"))
        except Exception as e:
            logger.error(f"Error updating attendance: {str(e)}")
            flash("An error occurred while updating attendance", "error")
            return redirect(url_for("show_attendance"))
    else:
        flash("Please log in to update attendance", "error")
        return redirect(url_for("login"))

@app.route("/delete_attendance/<int:id>", methods=["POST"])
def delete_attendance(id):
    if "username" in session:
        try:
            wb = openpyxl.load_workbook(attendance_file)
            sheet = wb.active
            
            for row in sheet.iter_rows(min_row=2):
                if row[0].value == id:
                    sheet.delete_rows(row[0].row)
                    break
            
            wb.save(attendance_file)
            flash("Attendance deleted successfully!", "success")
            return redirect(url_for("show_attendance"))
        except Exception as e:
            logger.error(f"Error deleting attendance: {str(e)}")
            flash("An error occurred while deleting attendance", "error")
            return redirect(url_for("show_attendance"))
    else:
        flash("Please log in to delete attendance", "error")
        return redirect(url_for("login"))

if __name__ == "__main__":
    port = int(os.environ.get("PORT", 10000))
    app.run(host="0.0.0.0", port=port)
